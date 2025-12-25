"""
Fonctions utilitaires et helpers techniques.

Fichier: cabinet/utils/export.py
"""

# ==================== Imports ====================
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import os
from utilisateurs.models import Client 
from reclamations.models import Reclamation
from django.http import HttpResponse
from fiscal.models import CMIR , SuiviForfaitaire
from juridique.models import JuridiqueCreation, DocumentJuridique, EvenementJuridique
# Ajoute MSYS2 aux variables d'environnement
os.environ['PATH'] = r'C:\msys64\mingw64\bin;' + os.environ['PATH']
os.environ['FONTCONFIG_PATH'] = r'C:\msys64\mingw64\etc\fonts'


from comptables.models import Comptable
from dossiers.models import Dossier
from honoraires.models import Honoraire , HonorairePV 
from honoraires.models import ReglementHonorairePV
from fiscal.models import Acompte , DepotBilan



# ==================== Classes ====================
class ExportExcel:
    """Classe pour gérer les exports Excel"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Configure les styles pour l'export"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def apply_header_style(self, worksheet, row_num, max_col):
        """Applique le style d'en-tête à une ligne"""
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
    
    def auto_adjust_columns(self, worksheet):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_comptables(self):
        """Exporte la liste des comptables"""
        worksheet = self.workbook.active
        worksheet.title = "Comptables"
        
        # En-têtes
        headers = [
            'Matricule', 'Nom', 'Prénom', 'CNSS', 'Téléphone', 'Email',
            'Email Personnel', 'Nb Dossiers PM', 'Nb Dossiers ET', 'Actif'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        self.apply_header_style(worksheet, 1, len(headers))
        
        # Données
        comptables = Comptable.objects.all().order_by('nom', 'prenom')
        for row, comptable in enumerate(comptables, 2):
            worksheet.cell(row=row, column=1, value=comptable.matricule)
            worksheet.cell(row=row, column=2, value=comptable.nom)
            worksheet.cell(row=row, column=3, value=comptable.prenom)
            worksheet.cell(row=row, column=4, value=comptable.cnss)
            worksheet.cell(row=row, column=5, value=comptable.tel)
            worksheet.cell(row=row, column=6, value=comptable.email)
            worksheet.cell(row=row, column=7, value=comptable.email_personnel)
            worksheet.cell(row=row, column=8, value=comptable.nbre_pm)
            worksheet.cell(row=row, column=9, value=comptable.nbre_et)
            worksheet.cell(row=row, column=10, value="Oui" if comptable.actif else "Non")
        
        self.auto_adjust_columns(worksheet)
        
        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="comptables_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def export_clients(self):
        """Exporte la liste des clients"""
        worksheet = self.workbook.active
        worksheet.title = "Clients"
        
        # En-têtes
        headers = [
            'Nom complet', 
            'Email', 
            'Téléphone', 
            'Nom d\'utilisateur',
            'Date création'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        self.apply_header_style(worksheet, 1, len(headers))
        
        # Données
        clients = Client.objects.all().order_by('user__last_name', 'user__first_name')
        for row, client in enumerate(clients, 2):
            worksheet.cell(row=row, column=1, value=f"{client.user.first_name} {client.user.last_name}" if client.user else client.contact_personne)
            worksheet.cell(row=row, column=2, value=client.email)
            worksheet.cell(row=row, column=3, value=client.user.telephone if client.user else "")
            worksheet.cell(row=row, column=4, value=client.user.username if client.user else "")
            worksheet.cell(row=row, column=5, value=client.date_creation.strftime("%d/%m/%Y") if client.date_creation else "")
        
        self.auto_adjust_columns(worksheet)
        
        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="clients_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response

    def export_acompte(self):
        """Exporte la liste des acomptes"""
        worksheet = self.workbook.active
        worksheet.title = "Acomptes"

        # En-têtes
        headers = [
            'Dossier', 'Comptable', 'Année', 'Montant IS',
            'Régularisation', 'A1', 'A2', 'A3', 'A4', 
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        acomptes = Acompte.objects.select_related('dossier', 'dossier__comptable_traitant').all()
        for row, acompte in enumerate(acomptes, 2):
            worksheet.cell(row=row, column=1, value=acompte.dossier.denomination if acompte.dossier else 'N/A')
            worksheet.cell(row=row, column=2, value=acompte.dossier.comptable_traitant.nom_complet if acompte.dossier and acompte.dossier.comptable_traitant else 'N/A')
            worksheet.cell(row=row, column=3, value=acompte.annee)
            worksheet.cell(row=row, column=4, value=float(acompte.is_montant))
            worksheet.cell(row=row, column=5, value=float(acompte.regularisation))
            worksheet.cell(row=row, column=6, value= acompte.a1 )
            worksheet.cell(row=row, column=7, value= acompte.a2 )
            worksheet.cell(row=row, column=8, value= acompte.a3 )
            worksheet.cell(row=row, column=9, value=acompte.a4)
            

        self.auto_adjust_columns(worksheet)

        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="acomptes_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response
    
    def export_dossiers(self):
        """Exporte la liste des dossiers"""
        worksheet = self.workbook.active
        worksheet.title = "Dossiers"
        
        # En-têtes
        headers = [
            'Code', 'Dénomination', 'Forme Juridique', 'Date Création',
            'Secteur Activités', 'Adresse', 'Email', 'Téléphone',
            'ICE', 'RC', 'Déclaration TVA', 'Comptable Traitant', 'Statut'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        self.apply_header_style(worksheet, 1, len(headers))
        
        # Données
        dossiers = Dossier.objects.select_related('comptable_traitant').filter(actif=True).order_by('denomination')
        for row, dossier in enumerate(dossiers, 2):
            worksheet.cell(row=row, column=1, value=dossier.code)
            worksheet.cell(row=row, column=2, value=dossier.denomination)
            worksheet.cell(row=row, column=3, value=dossier.get_forme_juridique_display())
            worksheet.cell(row=row, column=4, value=dossier.date_creation)
            worksheet.cell(row=row, column=5, value=dossier.secteur_activites)
            worksheet.cell(row=row, column=6, value=dossier.adresse)
            worksheet.cell(row=row, column=7, value=dossier.email)
            worksheet.cell(row=row, column=8, value=dossier.fixe or dossier.gsm)
            worksheet.cell(row=row, column=9, value=dossier.ice)
            worksheet.cell(row=row, column=10, value=dossier.rc)
            worksheet.cell(row=row, column=11, value=dossier.get_declaration_tva_display())
            worksheet.cell(row=row, column=12, value=dossier.comptable_traitant.nom_complet)
            worksheet.cell(row=row, column=13, value=dossier.get_statut_display())
        
        self.auto_adjust_columns(worksheet)
        
        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dossiers_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def export_honoraires(self):
        """Exporte la liste des honoraires avec règlements"""
        worksheet = self.workbook.active
        worksheet.title = "Honoraires"
        
        # En-têtes
        headers = [
            'Code','Dossier', 'Comptable', 'Montant Mensuel', 'Montant Trimestriel',

            'Montant Annuel', 'Total Réglé Mensuel', 'Total Réglé Trimestriel',
            'Total Réglé Annuel', 'Reste Mensuel', 'Reste Trimestriel', 'Reste Annuel'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        self.apply_header_style(worksheet, 1, len(headers))
        
        # Données
        honoraires = Honoraire.objects.select_related('dossier', 'dossier__comptable_traitant').all()
        for row, honoraire in enumerate(honoraires, 2):
            worksheet.cell(row=row, column=1, value=honoraire.code)
            worksheet.cell(row=row, column=2, value=honoraire.dossier.denomination)
            worksheet.cell(row=row, column=3, value=honoraire.dossier.comptable_traitant.nom_complet)
            worksheet.cell(row=row, column=4, value=float(honoraire.montant_mensuel))
            worksheet.cell(row=row, column=5, value=float(honoraire.montant_trimestriel))
            worksheet.cell(row=row, column=6, value=float(honoraire.montant_annuel))
            worksheet.cell(row=row, column=7, value=float(honoraire.total_reglements_mensuel))
            worksheet.cell(row=row, column=8, value=float(honoraire.total_reglements_trimestriel))
            worksheet.cell(row=row, column=9, value=float(honoraire.total_reglements_annuel))
            worksheet.cell(row=row, column=10, value=float(honoraire.reste_mensuel))
            worksheet.cell(row=row, column=11, value=float(honoraire.reste_trimestriel))
            worksheet.cell(row=row, column=12, value=float(honoraire.reste_annuel))
        
        self.auto_adjust_columns(worksheet)
        
        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="honoraires_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def export_honoraires_pv(self):
        """Exporte la liste des honoraires PV avec règlements"""
        worksheet = self.workbook.active
        worksheet.title = "Honoraires PV"

        # En-têtes
        headers = [
            'Code','Dossier', 'Comptable', 'Montant PV', 'Date PV',
            'Total Réglé PV', 'Reste PV'
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        honoraires_pv = HonorairePV.objects.select_related('dossier', 'dossier__comptable_traitant').all()
        for row, honoraire_pv in enumerate(honoraires_pv, 2):
            worksheet.cell(row=row, column=1,value=honoraire_pv.code)
            worksheet.cell(row=row,column=2,value=honoraire_pv.dossier.denomination if honoraire_pv.dossier else 'N/A')
            worksheet.cell(row=row,column=3,value=honoraire_pv.dossier.comptable_traitant.nom_complet if honoraire_pv.dossier and honoraire_pv.dossier.comptable_traitant else 'N/A')
            worksheet.cell(row=row, column=4, value=float(honoraire_pv.montant_total))
            worksheet.cell(row=row, column=5, value=honoraire_pv.date_creation.strftime("%Y-%m-%d") if honoraire_pv.date_creation else '')
            worksheet.cell(row=row, column=6, value=float(honoraire_pv.total_reglements_pv))
            worksheet.cell(row=row, column=7, value=float(honoraire_pv.reste_pv))

        self.auto_adjust_columns(worksheet)

        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="honoraires_pv_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    
    def export_suivi_tva(self, annee=None):
        """Exporte le suivi TVA pour une année donnée"""
        if not annee:
            annee = datetime.now().year
        
        worksheet = self.workbook.active
        worksheet.title = f"Suivi TVA {annee}"
        
        # En-têtes
        headers = [
            'Code','Dossier', 'Comptable', 'Type TVA',
            'Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun',
            'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc',
            'T1', 'T2', 'T3', 'T4', '% Mensuel', '% Trimestriel'
        ]
        
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        self.apply_header_style(worksheet, 1, len(headers))
        
        # Données
        from fiscal.models import SuiviTVA
        suivis = SuiviTVA.objects.select_related('dossier', 'dossier__comptable_traitant').filter(annee=annee)
        
        for row, suivi in enumerate(suivis, 2):
            worksheet.cell(row=row, column=1, value=suivi.code)
            worksheet.cell(row=row, column=2, value=suivi.dossier.denomination)
            worksheet.cell(row=row, column=3, value=suivi.dossier.comptable_traitant.nom_complet)
            worksheet.cell(row=row, column=4, value=suivi.dossier.get_declaration_tva_display())
            
            # Mois
            mois_values = [suivi.jan, suivi.fev, suivi.mars, suivi.avril, suivi.mai, suivi.juin,
                          suivi.juillet, suivi.aout, suivi.sep, suivi.oct, suivi.nov, suivi.dec]
            for col, value in enumerate(mois_values, 4):
                worksheet.cell(row=row, column=col, value="✓" if value else "")
            
            # Trimestres
            trimestres = [suivi.t1, suivi.t2, suivi.t3, suivi.t4]
            for col, value in enumerate(trimestres, 16):
                worksheet.cell(row=row, column=col, value="✓" if value else "")
            
            # Pourcentages
            worksheet.cell(row=row, column=20, value=f"{suivi.pourcentage_mensuel_complete:.1f}%")
            worksheet.cell(row=row, column=21, value=f"{suivi.pourcentage_trimestriel_complete:.1f}%")
        
        self.auto_adjust_columns(worksheet)
        
        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="suivi_tva_{annee}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response

    def export_cmir(self):
        """Exporte la liste des CM/IR"""
        worksheet = self.workbook.active
        worksheet.title = "CM_IR"

        # En-têtes
        headers = [
            'Code','Dossier', 'Comptable', 'Année',
            'Montant IR', 'Montant CM', 'Complément IR',
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        cm_ir_records = CMIR.objects.select_related('dossier', 'dossier__comptable_traitant').all()

        for row, record in enumerate(cm_ir_records, 2):
            worksheet.cell(row=row, column=1, value=record.code)
            worksheet.cell(row=row, column=2, value=record.dossier.denomination if record.dossier else 'N/A')
            worksheet.cell(row=row, column=3, value=record.dossier.comptable_traitant.nom_complet if record.dossier and record.dossier.comptable_traitant else 'N/A')
            worksheet.cell(row=row, column=4, value=record.annee)
            worksheet.cell(row=row, column=5, value=float(record.montant_ir))
            worksheet.cell(row=row, column=6, value=float(record.montant_cm))
            worksheet.cell(row=row, column=7, value=float(record.complement_ir))

        self.auto_adjust_columns(worksheet)

        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cm_ir_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    def export_depotbilan(self):
        """Exporte la liste des dépôts de bilan"""
        worksheet = self.workbook.active
        worksheet.title = "Dépôts Bilan"

        # En-têtes
        headers = [
            'Code','Dossier', 'Comptable', 'Année exercice', 'Date de dépôt',
            'Remarque', 'Statut'
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        depots = DepotBilan.objects.select_related('dossier', 'dossier__comptable_traitant').all()

        for row, depot in enumerate(depots, 2):
            worksheet.cell(row=row, column=1, value=depot.code)
            worksheet.cell(row=row, column=2, value=depot.dossier.denomination if depot.dossier else 'N/A')
            worksheet.cell(row=row, column=3, value=depot.dossier.comptable_traitant.nom_complet if depot.dossier and depot.dossier.comptable_traitant else 'N/A')
            worksheet.cell(row=row, column=4, value=depot.annee_exercice)
            worksheet.cell(row=row, column=5, value=depot.date_depot.strftime('%Y-%m-%d') if depot.date_depot else '')
            worksheet.cell(row=row, column=6, value=depot.remarque or '')
            worksheet.cell(row=row, column=7, value=dict(DepotBilan.STATUT_CHOICES).get(depot.statut, depot.statut))
            

        self.auto_adjust_columns(worksheet)

        # Préparer la réponse
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="depots_bilan_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    def export_suiviforfaitaire(self):
        """Exporte la liste des Suivis Forfaitaires"""
        worksheet = self.workbook.active
        worksheet.title = "Suivis Forfaitaires"

        # Ligne 1 : En-têtes multi-lignes
        headers_row_1 = [
            "CODE", "PP", "DÉCLARATION ANNUELLE", "", "PAIEMENT ANNUEL",
            "PAIEMENT TRIMESTRIEL", "", "", ""
        ]
        headers_row_2 = [
            "", "", "I.R", "9421", "", "1er Acompte", "2e Acompte",
            "3e Acompte", "4e Acompte", ""
        ]

        for col, header in enumerate(headers_row_1, 1):
            worksheet.cell(row=1, column=col, value=header)
        for col, header in enumerate(headers_row_2, 1):
            worksheet.cell(row=2, column=col, value=header)

        # Fusion des cellules pour titres
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # CODE
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)  # PP
        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)  # DÉCLARATION ANNUELLE
        worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)  # PAIEMENT ANNUEL
        worksheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)  # PAIEMENT TRIMESTRIEL
        

        self.apply_header_style(worksheet, 1, 9)
        self.apply_header_style(worksheet, 2, 9)

        # Données
        suivis = SuiviForfaitaire.objects.all()

        for row_idx, suivi in enumerate(suivis, start=3):
            worksheet.cell(row=row_idx, column=1, value=suivi.code)
            worksheet.cell(row=row_idx, column=2, value=suivi.pp)
            worksheet.cell(row=row_idx, column=3, value=float(suivi.ir))
            worksheet.cell(row=row_idx, column=4, value=float(suivi.code_9421))
            worksheet.cell(row=row_idx, column=5, value=float(suivi.paiement_annuel))  # propriété calculée
            worksheet.cell(row=row_idx, column=6, value=float(suivi.acompte_1))
            worksheet.cell(row=row_idx, column=7, value=float(suivi.acompte_2))
            worksheet.cell(row=row_idx, column=8, value=float(suivi.acompte_3))
            worksheet.cell(row=row_idx, column=9, value=float(suivi.acompte_4))
            

        self.auto_adjust_columns(worksheet)

        # Export Excel
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="suivis_forfaitaires_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    def export_juridique_creation(self):
        """Exporte la liste des créations juridiques"""
        worksheet = self.workbook.active
        worksheet.title = "Créations Juridiques"

        # En-têtes
        headers = [
            "Code","Dossier", "Ordre", "Forme",
            "Certificat Négatif", "Date Certificat",
            "Statuts", "Date Statuts",
            "Contrat de Bail", "Date Contrat de Bail",
            "TP", "Date TP",
            "RC", "Date RC", "Numéro RC",
            "CNSS", "Date CNSS", "Numéro CNSS",
            "AL", "Date AL",
            "BO", "Date BO",
            "RQ", "Date RQ",
            "Statut Global", "Remarques",
            "Complétion (%)"
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        creations = JuridiqueCreation.objects.select_related('dossier').all()

        for row_idx, jc in enumerate(creations, start=2):
            worksheet.cell(row=row_idx, column=1, value=jc.code)
            worksheet.cell(row=row_idx, column=2, value=jc.dossier.denomination if jc.dossier else 'N/A')
            worksheet.cell(row=row_idx, column=3, value=jc.ordre)
            worksheet.cell(row=row_idx, column=4, value=jc.forme)

            worksheet.cell(row=row_idx, column=5, value="Oui" if jc.certificat_negatif else "Non")
            worksheet.cell(row=row_idx, column=6, value=jc.date_certificat_negatif)

            worksheet.cell(row=row_idx, column=7, value="Oui" if jc.statuts else "Non")
            worksheet.cell(row=row_idx, column=8, value=jc.date_statuts)

            worksheet.cell(row=row_idx, column=9, value="Oui" if jc.contrat_bail else "Non")
            worksheet.cell(row=row_idx, column=10, value=jc.date_contrat_bail)

            worksheet.cell(row=row_idx, column=11, value="Oui" if jc.tp else "Non")
            worksheet.cell(row=row_idx, column=12, value=jc.date_tp)

            worksheet.cell(row=row_idx, column=13, value="Oui" if jc.rc else "Non")
            worksheet.cell(row=row_idx, column=14, value=jc.date_rc)
            worksheet.cell(row=row_idx, column=15, value=jc.numero_rc)

            worksheet.cell(row=row_idx, column=16, value="Oui" if jc.cnss else "Non")
            worksheet.cell(row=row_idx, column=17, value=jc.date_cnss)
            worksheet.cell(row=row_idx, column=18, value=jc.numero_cnss)

            worksheet.cell(row=row_idx, column=19, value="Oui" if jc.al else "Non")
            worksheet.cell(row=row_idx, column=20, value=jc.date_al)

            worksheet.cell(row=row_idx, column=21, value="Oui" if jc.bo else "Non")
            worksheet.cell(row=row_idx, column=22, value=jc.date_bo)

            worksheet.cell(row=row_idx, column=23, value="Oui" if jc.rq else "Non")
            worksheet.cell(row=row_idx, column=24, value=jc.date_rq)

            worksheet.cell(row=row_idx, column=25, value=jc.get_statut_global_display())
            worksheet.cell(row=row_idx, column=26, value=jc.remarques)
            worksheet.cell(row=row_idx, column=27, value=round(jc.pourcentage_completion, 2))

        self.auto_adjust_columns(worksheet)

        # Export Excel
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="juridique_creation_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    def export_document_juridique(self):
        """Exporte les documents juridiques au format Excel"""
        worksheet = self.workbook.active
        worksheet.title = "Documents Juridiques"

        # En-têtes
        headers = [
            "Code",
            "Dossier",
            "Nom du document",
            "Type de document",
            "Description",
            "Date du document",
            "Statut",
            "Lien fichier",
            
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        documents = DocumentJuridique.objects.select_related('dossier').all()

        for row_idx, doc in enumerate(documents, start=2):
            worksheet.cell(row=row_idx, column=1, value=doc.code)
            worksheet.cell(row=row_idx, column=2, value=doc.dossier.denomination if doc.dossier else "N/A")
            worksheet.cell(row=row_idx, column=3, value=doc.nom_document)
            worksheet.cell(row=row_idx, column=4, value=doc.get_type_document_display())
            worksheet.cell(row=row_idx, column=5, value=doc.description)
            worksheet.cell(row=row_idx, column=6, value=doc.date_document)
            worksheet.cell(row=row_idx, column=7, value=doc.get_statut_display())
            
            # Générer le lien du fichier (chemin relatif ou complet selon usage)
            if doc.fichier:
                worksheet.cell(row=row_idx, column=8, value=doc.fichier.url)
            else:
                worksheet.cell(row=row_idx, column=8, value="Aucun fichier")

           

        self.auto_adjust_columns(worksheet)

        # Export Excel
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="documents_juridiques_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    def export_evenement_juridique(self):
        """Exporte les événements juridiques au format Excel"""
        worksheet = self.workbook.active
        worksheet.title = "Événements Juridiques"

        # En-têtes
        headers = [
            "Code",
            "Dossier",
            "Type d'événement",
            "Description",
            "Date de l'événement",
            "Statut",
            
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        evenements = EvenementJuridique.objects.select_related('dossier').all()

        for row_idx, evt in enumerate(evenements, start=2):
            worksheet.cell(row=row_idx, column=1, value=evt.code)
            worksheet.cell(row=row_idx, column=2, value=evt.dossier.denomination if evt.dossier else "N/A")
            worksheet.cell(row=row_idx, column=3, value=evt.get_type_evenement_display())
            worksheet.cell(row=row_idx, column=4, value=evt.description)
            worksheet.cell(row=row_idx, column=5, value=evt.date_evenement.strftime("%Y-%m-%d") if evt.date_evenement else "")
            worksheet.cell(row=row_idx, column=6, value=evt.get_statut_display())
            

        self.auto_adjust_columns(worksheet)

        # Générer le fichier Excel
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="evenements_juridiques_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response
    
    def export_reclamations(self):
        """Exporte les réclamations au format Excel"""
        worksheet = self.workbook.active
        worksheet.title = "Réclamations"

        # En-têtes
        headers = [
            "Code",
            "Dossier"
            "Type de réclamation",
            "Date de réception",
            "Réponse",
            "Suivi",
            "Remarque",
            "Destinataire",
            "Créée par",
            "Classement",
            "Priorité"
        
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)

        self.apply_header_style(worksheet, 1, len(headers))

        # Données
        reclamations = Reclamation.objects.select_related('dossier', 'created_by', 'destinataire').all()

        for row_idx, r in enumerate(reclamations, start=2):
            worksheet.cell(row=row_idx, column=1, value=r.code)
            worksheet.cell(row=row_idx, column=2, value=r.dossier.denomination if r.dossier else "")
            worksheet.cell(row=row_idx, column=3, value=r.type_reclamation)
            worksheet.cell(row=row_idx, column=4, value=r.date_reception.strftime("%Y-%m-%d") if r.date_reception else "")
            worksheet.cell(row=row_idx, column=5, value=r.reponse)
            worksheet.cell(row=row_idx, column=6, value=r.suivi)
            worksheet.cell(row=row_idx, column=7, value=r.remarque)
            worksheet.cell(row=row_idx, column=8, value=str(r.destinataire) if r.destinataire else "")
            worksheet.cell(row=row_idx, column=9, value=str(r.created_by))
            worksheet.cell(row=row_idx, column=10, value=r.get_classement_display())
            worksheet.cell(row=row_idx, column=11, value=r.get_priorite_display())
        

        self.auto_adjust_columns(worksheet)

        # Générer le fichier Excel
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reclamations_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response